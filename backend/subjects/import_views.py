import io
import json as _json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, JSONParser
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from django.http import HttpResponse

from .models import Subject, Topic, Question, Answer
from .permissions import IsTeacherOrAdmin
from .google_forms.parser import ParseError, parse_dict, raw_to_payload
from .google_forms.importer import ImporterError, import_payload
from .google_forms.fetcher import FetcherError, fetch_form
from .google_forms.answers import is_available as forms_api_available, get_form_quiz_data, merge_into_questions


class ExcelTemplateView(APIView):
    """Скачивание пустого шаблона Excel для импорта вопросов"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Вопросы'

        headers = ['Тема', 'Текст вопроса', 'Сложность', 'Вариант A', 'Вариант B',
                   'Вариант C', 'Вариант D', 'Правильный ответ', 'Объяснение']
        ws.append(headers)

        # Стилизация заголовков
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Пример строки
        example = ['Алгебра', 'Чему равно 2 + 2?', 'easy', '3', '4', '5', '6', 'B', '2 + 2 = 4']
        ws.append(example)

        # Инструкция
        ws2 = wb.create_sheet('Инструкция')
        instructions = [
            ['ИНСТРУКЦИЯ ПО ЗАПОЛНЕНИЮ', ''],
            ['', ''],
            ['1. Тема', 'Название темы (если темы нет — она создастся автоматически)'],
            ['2. Текст вопроса', 'Сам вопрос'],
            ['3. Сложность', 'easy (лёгкий), medium (средний), hard (сложный)'],
            ['4-7. Варианты', '4 варианта ответа (A, B, C, D)'],
            ['8. Правильный ответ', 'Буква правильного ответа: A, B, C или D'],
            ['9. Объяснение', 'Объяснение правильного ответа (необязательно)'],
            ['', ''],
            ['ВАЖНО:', ''],
            ['', '• Не удаляйте строку с заголовками'],
            ['', '• Не меняйте порядок столбцов'],
            ['', '• Правильный ответ — только одна буква: A, B, C или D'],
        ]
        for row in instructions:
            ws2.append(row)
        ws2['A1'].font = Font(bold=True, size=14)

        # Авто-ширина
        for ws_sheet in [ws, ws2]:
            for column in ws_sheet.columns:
                max_length = max(len(str(cell.value or '')) for cell in column)
                ws_sheet.column_dimensions[column[0].column_letter].width = min(max_length + 4, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="ent_questions_template.xlsx"'
        return response


class ExcelImportView(APIView):
    """Импорт вопросов из Excel файла"""
    permission_classes = [IsTeacherOrAdmin]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get('file')
        subject_id = request.data.get('subject_id')

        if not file:
            return Response({'error': 'Файл не загружен'}, status=status.HTTP_400_BAD_REQUEST)
        if not subject_id:
            return Response({'error': 'Не выбран предмет'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            subject = Subject.objects.get(id=subject_id)
        except Subject.DoesNotExist:
            return Response({'error': 'Предмет не найден'}, status=status.HTTP_404_NOT_FOUND)

        try:
            wb = Workbook()
            wb = wb.load_workbook(file, data_only=True)
        except Exception:
            return Response({'error': 'Не удалось прочитать Excel-файл. Убедитесь, что это .xlsx'}, status=400)

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        created = 0
        errors = []
        difficulty_map = {'easy': 'easy', 'medium': 'medium', 'hard': 'hard',
                          'лёгкий': 'easy', 'легкий': 'easy', 'средний': 'medium', 'сложный': 'hard'}

        for idx, row in enumerate(rows, start=2):
            # Пропускаем пустые строки
            if not row or not any(str(c).strip() for c in row if c is not None):
                continue

            try:
                # Распаковываем с учётом возможных пустых ячеек
                topic_name = str(row[0] or '').strip()
                question_text = str(row[1] or '').strip()
                difficulty_raw = str(row[2] or 'medium').strip().lower()
                ans_a = str(row[3] or '').strip()
                ans_b = str(row[4] or '').strip()
                ans_c = str(row[5] or '').strip()
                ans_d = str(row[6] or '').strip()
                correct_raw = str(row[7] or '').strip().upper()
                explanation = str(row[8] or '').strip()

                # Валидация
                if not question_text:
                    errors.append(f'Строка {idx}: пустой текст вопроса')
                    continue
                if not topic_name:
                    errors.append(f'Строка {idx}: не указана тема')
                    continue
                if correct_raw not in ('A', 'B', 'C', 'D'):
                    errors.append(f'Строка {idx}: правильный ответ должен быть A, B, C или D (получено "{correct_raw}")')
                    continue

                answers_texts = {'A': ans_a, 'B': ans_b, 'C': ans_c, 'D': ans_d}
                if not answers_texts[correct_raw]:
                    errors.append(f'Строка {idx}: не заполнен правильный вариант ({correct_raw})')
                    continue

                difficulty = difficulty_map.get(difficulty_raw, 'medium')

                # Создаём или берём тему
                topic, _ = Topic.objects.get_or_create(
                    name=topic_name, subject=subject
                )

                # Создаём вопрос
                question = Question.objects.create(
                    text=question_text,
                    topic=topic,
                    difficulty=difficulty,
                    question_type='single_choice',
                    source_type='teacher_upload',
                    points=1,
                    explanation=explanation,
                )

                # Создаём ответы
                for letter, text in answers_texts.items():
                    if text:  # создаём только непустые варианты
                        Answer.objects.create(
                            question=question,
                            text=text,
                            is_correct=(letter == correct_raw),
                        )

                created += 1

            except Exception as e:
                errors.append(f'Строка {idx}: {str(e)}')
                continue

        return Response({
            'created': created,
            'errors': errors,
            'errors_count': len(errors),
            'message': f'Импортировано {created} вопросов' + (f', ошибок: {len(errors)}' if errors else ''),
        }, status=status.HTTP_200_OK)


class GoogleFormsImportView(APIView):
    """
    Импорт авторских тестов из JSON-экспорта Google Forms.

    Принимает JSON двумя способами:
      1. POST application/json           — тело запроса = экспорт Apps Script.
      2. POST multipart/form-data        — файл .json в поле ``file``.

    Query-параметры:
      ?dry_run=1                         — только валидация, без записи в БД.

    Доступ: teacher / admin (см. IsTeacherOrAdmin).
    """
    permission_classes = [IsTeacherOrAdmin]
    parser_classes = [JSONParser, MultiPartParser]

    def post(self, request):
        payload_dict, source_label = self._extract_payload(request)
        if payload_dict is None:
            return Response(
                {'error': source_label},
                status=status.HTTP_400_BAD_REQUEST,
            )

        dry_run = self._truthy(request.query_params.get('dry_run'))

        # 1. Парсинг + валидация структуры.
        try:
            payload = parse_dict(payload_dict)
        except ParseError as e:
            return Response(
                {'error': f'Ошибка разбора JSON: {e}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        meta = {
            'form_title': payload.form_title,
            'form_id': payload.form_id,
            'subject': payload.subject_slug or payload.subject_name,
            'topic': payload.topic_name,
            'language': payload.language,
            'year': payload.year,
            'questions_in_file': len(payload.questions),
            'source': source_label,
            'dry_run': dry_run,
        }

        # 2. Сухой прогон — отчёт без записи.
        if dry_run:
            return Response(
                {'ok': True, 'dry_run': True, 'meta': meta},
                status=status.HTTP_200_OK,
            )

        # 3. Реальный импорт.
        try:
            result = import_payload(payload)
        except ImporterError as e:
            # Предмет не найден и т.п. — пользовательская ошибка.
            return Response(
                {'error': str(e), 'meta': meta},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(
            {'ok': True, 'result': result.as_dict(), 'meta': meta},
            status=status.HTTP_201_CREATED,
        )

    # --- helpers -----------------------------------------------------------

    def _extract_payload(self, request):
        """
        Возвращает (dict, label) где dict — JSON-экспорт, label — описание
        источника для отчёта. При ошибке возвращает (None, error_message).
        """
        # Вариант 1: multipart с полем file.
        f = request.FILES.get('file')
        if f is not None:
            if not f.name.lower().endswith('.json'):
                return None, 'Ожидался .json файл'
            try:
                raw = f.read().decode('utf-8')
                return _json.loads(raw), f'file:{f.name}'
            except (UnicodeDecodeError, _json.JSONDecodeError) as e:
                return None, f'Невалидный JSON в файле: {e}'

        # Вариант 2: application/json — DRF уже распарсил в request.data.
        data = request.data
        if isinstance(data, dict):
            return data, 'json-body'

        return None, 'Тело запроса должно быть JSON-объектом или .json файлом'

    @staticmethod
    def _truthy(value) -> bool:
        if value is None:
            return False
        return str(value).strip().lower() in {'1', 'true', 'yes', 'on'}


def _resolve_form_id(value: str) -> tuple[str, str]:
    """
    Извлекает form_id и form_url из переданного значения.

    Если передан URL — возвращает (form_id, URL).
    Если передан голый ID — возвращает (ID, '') (скрапинг работать не будет).
    """
    from .google_forms.fetcher import _extract_form_id
    form_id = _extract_form_id(value)
    if form_id:
        return form_id, value
    # Голый ID — используем как есть
    return value, ''


def import_from_google_form(
    form_id: str,
    subject_slug: str = '',
    subject_name: str = '',
    topic_name: str = '',
    language: str = 'ru',
    year: int | None = None,
    dry_run: bool = False,
    service_account_email: str | None = None,
) -> dict:
    """
    Общая логика импорта Google Form (URL или ID).

    Возвращает dict с ключами:
      ok, error?, result?, meta, attached_images?
    """
    api_form_id, form_url = _resolve_form_id(form_id)
    source = f'id:{api_form_id}'

    # 1. Скрапинг (только если есть form_url)
    if form_url:
        try:
            fetch_result = fetch_form(form_url)
        except FetcherError as e:
            return {'ok': False, 'error': str(e)}
        source = f'url:{form_url[:80]}'
    else:
        from .google_forms.fetcher import FetchResult
        fetch_result = FetchResult()
        fetch_result.form_id = api_form_id

    # 2. Правильные ответы и баллы через Forms API
    if api_form_id and forms_api_available():
        try:
            quiz_data = get_form_quiz_data(api_form_id)
            if quiz_data:
                fetch_result.questions = merge_into_questions(
                    fetch_result.questions,
                    quiz_data.get('answers', []),
                    quiz_data.get('questions', []),
                )
                fetch_result.used_api = True
                fetch_result.warnings.append(
                    f'Получено {len(quiz_data.get("answers", []))} правильных ответов из Forms API'
                )
        except Exception:
            fetch_result.warnings.append('Ошибка получения правильных ответов из Forms API')

    # 3. Картинки — сначала Forms API (прямые URL), потом Drive (fallback)
    from .google_forms.answers import get_form_images_via_api
    from .google_forms.images import download_batch

    if api_form_id:
        try:
            api_images = get_form_images_via_api(api_form_id)
            if api_images:
                fetch_result.images.update(api_images)
                fetch_result.warnings.append(
                    f'Скачано {len(api_images)} картинок через Forms API'
                )
        except Exception:
            fetch_result.warnings.append('Ошибка получения картинок из Forms API')

    if not fetch_result.images and fetch_result.image_drive_ids:
        try:
            downloaded_map = download_batch(fetch_result.image_drive_ids)
            fetch_result.images.update(downloaded_map)
            count = len(downloaded_map)
            total = len(fetch_result.image_drive_ids)
            if count == total:
                fetch_result.warnings.append(f'Скачано {count} картинок из Drive')
            elif count > 0:
                fetch_result.warnings.append(f'Скачано {count} из {total} картинок из Drive')
            else:
                fetch_result.warnings.append('Не удалось скачать картинки из Drive')
        except Exception:
            fetch_result.warnings.append('Ошибка скачивания картинок из Drive')

    # 4. Конвертируем в FormPayload
    if not fetch_result.questions:
        return {
            'ok': False,
            'error': 'Нет вопросов для импорта. Укажите form_url для скрапинга '
                     'или проверьте form_id для Forms API.',
            'meta': {
                'used_api': fetch_result.used_api,
                'used_scrape': fetch_result.used_scrape,
                'warnings': fetch_result.warnings,
            }
        }

    try:
        payload = raw_to_payload(
            {
                'form_id': fetch_result.form_id or api_form_id,
                'form_title': fetch_result.form_title or f'Form {api_form_id}',
                'form_url': form_url or api_form_id,
                'questions': fetch_result.questions,
            },
            subject_slug=subject_slug,
            subject_name=subject_name,
            topic_name=topic_name,
            language=language,
            year=year,
        )
    except ParseError as e:
        return {'ok': False, 'error': f'Ошибка обработки: {e}'}

    # Мета
    q_count = len(payload.questions)
    correct_count = sum(1 for q in payload.questions if q.has_correct)
    images_count = sum(len(imgs) for imgs in fetch_result.images.values()) + len(fetch_result.unmatched_images)

    meta = {
        'form_title': fetch_result.form_title or api_form_id,
        'form_id': fetch_result.form_id or api_form_id,
        'subject': subject_slug or subject_name,
        'topic': topic_name,
        'language': language,
        'year': year,
        'questions_in_form': q_count,
        'with_correct_answers': correct_count,
        'with_images': images_count,
        'source': source,
        'used_api': fetch_result.used_api,
        'used_scrape': fetch_result.used_scrape,
        'warnings': fetch_result.warnings,
        'dry_run': dry_run,
    }

    if dry_run:
        return {'ok': True, 'dry_run': True, 'meta': meta}

    # Реальный импорт
    try:
        result = import_payload(payload)
    except ImporterError as e:
        return {'ok': False, 'error': str(e), 'meta': meta}

    attached_images = 0
    if fetch_result.images:
        attached_images = _attach_images(fetch_result, payload, result)

    return {
        'ok': True,
        'result': result.as_dict(),
        'meta': meta,
        'attached_images': attached_images,
    }


class GoogleFormsUrlImportView(APIView):
    """
    Импорт авторских тестов из Google Form.

    Принимает form_id (URL или голый ID редактора).

    Body (JSON):
      {
        "form_id": "...",                 // ID формы (URL или голый ID редактора)
        "subject_slug": "math",            // или subject_name
        "topic_name": "Законы Ньютона",
        "service_account_email": "...",    // опционально
        "language": "ru",                  // опционально
        "year": 2026                       // опционально
      }

    Query-параметры:
      ?dry_run=1 — только предпросмотр, без записи в БД.

    Доступ: teacher / admin.
    """
    permission_classes = [IsTeacherOrAdmin]
    parser_classes = [JSONParser]

    @staticmethod
    def _truthy(value) -> bool:
        if value is None:
            return False
        return str(value).strip().lower() in {'1', 'true', 'yes', 'on'}

    def post(self, request):
        raw = str(request.data.get('form_id', '')).strip()
        if not raw:
            return Response(
                {'error': 'Укажите form_id (ссылка на форму или ID из редактора)'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        subject_slug = str(request.data.get('subject_slug', '')).strip()
        subject_name = str(request.data.get('subject_name', '')).strip()
        topic_name = str(request.data.get('topic_name', '')).strip()
        language = str(request.data.get('language', 'ru')).strip() or 'ru'
        service_account_email = str(request.data.get('service_account_email', '')).strip() or None
        year = None
        if request.data.get('year'):
            try:
                year = int(request.data['year'])
            except (TypeError, ValueError):
                pass

        if not (subject_slug or subject_name):
            return Response(
                {'error': 'Укажите предмет (subject_slug или subject_name)'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not topic_name:
            return Response(
                {'error': 'Укажите тему (topic_name)'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        dry_run = self._truthy(request.query_params.get('dry_run'))

        result = import_from_google_form(
            form_id=raw,
            subject_slug=subject_slug,
            subject_name=subject_name,
            topic_name=topic_name,
            language=language,
            year=year,
            dry_run=dry_run,
            service_account_email=service_account_email,
        )

        if not result['ok']:
            status_code = status.HTTP_400_BAD_REQUEST
            if 'error' in result:
                return Response(result, status=status_code)
            return Response(result, status=status_code)

        if result.get('dry_run'):
            return Response(result, status=status.HTTP_200_OK)

        return Response(result, status=status.HTTP_201_CREATED)


# ── Хелперы для URL-импорта ──────────────────────────────────────────────────

def _attach_images(fetch_result, payload, import_result) -> int:
    """
    Привязывает скачанные картинки к только что созданным вопросам.

    Сопоставление по порядку: payload.questions[i] ↔ Question в БД.
    """
    from io import BytesIO
    from django.core.files.images import ImageFile
    from .google_forms.images import _detect_ext

    attached = 0
    # Получаем все только что созданные вопросы в правильном порядке
    questions = Question.objects.filter(
        topic__name=import_result.topic,
        source_type='authorial',
    ).order_by('-id')[:len(payload.questions)]

    # Сопоставляем по external_id
    q_by_extid = {}
    for q in questions:
        if q.external_id:
            q_by_extid[q.external_id] = q

    for i, q_dto in enumerate(payload.questions):
        imgs = fetch_result.images.get(i, [])
        if not imgs:
            continue
        question = q_by_extid.get(q_dto.external_id)
        if not question:
            continue
        img_bytes = imgs[0]
        try:
            ext = _detect_ext(img_bytes)
            filename = f'gf_{question.id}_{i}.{ext}'
            question.image.save(
                filename,
                ImageFile(BytesIO(img_bytes), name=filename),
                save=True,
            )
            question.image_ref = ''
            question.save(update_fields=['image_ref'])
            attached += 1
        except Exception:
            pass

    return attached
